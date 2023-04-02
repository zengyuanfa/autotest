
# 引入openpyxl库
from openpyxl import Workbook, load_workbook

from openpyxl import load_workbook


class ExcelUtil:
    def __init__(self, file_name):
        # 初始化工作簿
        self.workbook = load_workbook(file_name, data_only=True)
        self.sheet_names = self.workbook.sheetnames

    def read_cell(self, sheet_name, row, col):
        # 读取单元格数据
        sheet = self.workbook[sheet_name]
        return sheet.cell(row=row, column=col).value

    def write_cell(self, sheet_name, row, col, data):
        # 写入单元格数据
        sheet = self.workbook[sheet_name]
        sheet.cell(row=row, column=col).value = data

    def save(self):
        # 保存工作簿
        self.workbook.save(self.file_name)


if __name__ == "__main__":
    file_path = "file/excel.xlsx"
    excel_util = ExcelUtil(file_path)
    value = excel_util.read_cell('Sheet1', 6, 2)
    print(value == 13.1)  # 输出原始值
    # excel_util.set_cell_value(1, 1, 1.23)
    # excel_util.save()
